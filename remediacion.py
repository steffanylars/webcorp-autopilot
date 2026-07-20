"""
============================================================
WEBCORP AUTOPILOT — Artefactos de remediacion
============================================================
Genera los artefactos que cierran el ciclo "from system alerts to
automated remediation" (texto literal del Track 4), SOLO despues de
la aprobacion humana (se invoca desde notificar_operaciones, que ya
esta detras del checkpoint HITL).

Reglas deterministicas — las decide ESTE codigo, no el LLM:
  - Caso puntual (municipio + carrier)  -> 1 PDF con marca dirigido a
    esa mensajeria, + borrador de correo a la mensajeria.
  - Ordenes no entregadas con clientes  -> 1 Excel para call center,
    + borrador de correo a call center.
  - Ambos pueden coexistir (PDF responsabiliza al carrier; Excel
    recupera clientes) — las dos unicas areas soportadas por ahora.

Todo el contenido es deterministico: los numeros vienen del estimador
Wilson y del endpoint de ordenes, la accion sugerida viene del sistema
de alertas. El LLM no redacta ni calcula nada aqui.

i18n: los artefactos se generan en el idioma del panel (es|en|fr) —
el parametro `idioma` llega desde el boton de aprobacion del humano,
nunca del LLM.

El correo NUNCA se envia automaticamente: se genera el borrador (.txt)
y un link mailto: para abrirlo en el cliente de correo del humano.

PII: el Excel contiene telefonos de clientes (uso interno de call
center). El directorio de artefactos esta en .gitignore y nunca se
publica. La PII no pasa por el LLM en ningun momento.

Dependencias: weasyprint (PDF; en macOS requiere
DYLD_FALLBACK_LIBRARY_PATH=/opt/homebrew/lib para hallar pango; en
Ubuntu, libpango-1.0-0). Si weasyprint no esta disponible, cae a un
HTML imprimible con la misma plantilla — el flujo nunca se rompe.
"""
import os
import json
import urllib.parse
from datetime import datetime, timezone

EMAIL_CALLCENTER = os.getenv("EMAIL_CALLCENTER", "callcenter@webcorp.example")
# Dominio .example (RFC 2606): placeholder deliberado hasta tener los reales.
EMAIL_CARRIER_FMT = os.getenv("EMAIL_CARRIER_FMT", "operaciones@{carrier}.example")


def _dir_artefactos() -> str:
    d = os.getenv("ARTEFACTOS_DIR", os.path.join(os.path.dirname(__file__), "artefactos"))
    os.makedirs(d, exist_ok=True)
    return d


def _slug(s: str) -> str:
    return "".join(c if c.isalnum() else "-" for c in (s or "").lower()).strip("-")[:40]


# ------------------------------------------------------------------
# i18n de artefactos — el idioma lo elige el humano en el panel.
# Fallback: espanol. Deterministico: son plantillas, no redaccion LLM.
# ------------------------------------------------------------------
_L = {
    "es": {
        "pdf_footer": "WebCorp Autopilot · generado tras aprobación humana · ",
        "pdf_h1": "Solicitud de remediación operativa",
        "pdf_dirigida": "dirigida a",
        "pdf_caso": "Caso detectado",
        "titulo": "Título", "zona": "Zona", "mensajeria": "Mensajería",
        "muestra": "Muestra (n)", "cruda": "Tasa cruda",
        "wilson": "Efectividad Wilson LCB 95%", "snapshot": "Snapshot de datos",
        "no_entregadas": "Órdenes no entregadas detectadas",
        "accion": "Acción sugerida", "trazabilidad": "Trazabilidad",
        "pdf_meta": ("Metodología: ranking por límite inferior del intervalo de Wilson (95%), "
                     "nunca por tasa cruda · Este documento fue generado automáticamente por "
                     "WebCorp Autopilot <b>después de aprobación humana explícita</b>; los números "
                     "provienen de herramientas determinísticas, no de un modelo de lenguaje."),
        "xl_hoja": "Ordenes no entregadas",
        "xl_cols": ["Orden", "Teléfono", "Municipio", "Depto", "Carrier",
                    "Estatus", "Subestatus", "COD", "Fecha"],
        "cc_asunto": "[Autopilot] Gestión call center — {n} órdenes no entregadas · {muni} × {carrier}",
        "cc_extra": ("Se adjunta el Excel con {n} órdenes no entregadas "
                     "(orden, teléfono, estatus, COD) para gestión de recuperación."),
        "carrier_asunto": "[Autopilot] Remediación requerida — {carrier} en {muni}, {depto}",
        "carrier_extra": "Se adjunta el PDF con el detalle del caso y la metodología.",
        "sin_fila": "sin fila en el estimador",
        "saludo": "Estimado equipo,",
        "intro": "El sistema de monitoreo de WebCorp detectó y un operador humano aprobó escalar el siguiente caso:",
        "datos_caso": "Datos del caso (snapshot {snapshot}):",
        "evidencia": "Evidencia",
        "firma": ("Borrador generado por WebCorp Autopilot tras aprobación humana.\n"
                  "Este correo NO se envía automáticamente: requiere revisión y envío manual."),
        "para": "Para", "asunto_lbl": "Asunto",
        "nota_final": ("Artefactos deterministas generados tras aprobación humana. "
                       "Los correos son BORRADORES: nunca se envían automáticamente."),
        "accion_default": "Revisión operativa del par municipio-carrier.",
    },
    "en": {
        "pdf_footer": "WebCorp Autopilot · generated after human approval · ",
        "pdf_h1": "Operational remediation request",
        "pdf_dirigida": "addressed to",
        "pdf_caso": "Detected case",
        "titulo": "Title", "zona": "Zone", "mensajeria": "Carrier",
        "muestra": "Sample (n)", "cruda": "Raw rate",
        "wilson": "Wilson LCB 95% effectiveness", "snapshot": "Data snapshot",
        "no_entregadas": "Undelivered orders detected",
        "accion": "Suggested action", "trazabilidad": "Traceability",
        "pdf_meta": ("Methodology: ranking by the lower bound of the Wilson interval (95%), "
                     "never by raw rate · This document was generated automatically by "
                     "WebCorp Autopilot <b>after explicit human approval</b>; every figure "
                     "comes from deterministic tools, not from a language model."),
        "xl_hoja": "Undelivered orders",
        "xl_cols": ["Order", "Phone", "Municipality", "Department", "Carrier",
                    "Status", "Substatus", "COD", "Date"],
        "cc_asunto": "[Autopilot] Call-center follow-up — {n} undelivered orders · {muni} × {carrier}",
        "cc_extra": ("Attached is the Excel with {n} undelivered orders "
                     "(order, phone, status, COD) for recovery follow-up."),
        "carrier_asunto": "[Autopilot] Remediation required — {carrier} in {muni}, {depto}",
        "carrier_extra": "Attached is the PDF with the case detail and methodology.",
        "sin_fila": "no row in the estimator",
        "saludo": "Dear team,",
        "intro": "WebCorp's monitoring system detected — and a human operator approved escalating — the following case:",
        "datos_caso": "Case data (snapshot {snapshot}):",
        "evidencia": "Evidence",
        "firma": ("Draft generated by WebCorp Autopilot after human approval.\n"
                  "This email is NOT sent automatically: it requires review and manual sending."),
        "para": "To", "asunto_lbl": "Subject",
        "nota_final": ("Deterministic artifacts generated after human approval. "
                       "Emails are DRAFTS: they are never sent automatically."),
        "accion_default": "Operational review of the municipality-carrier pair.",
    },
    "fr": {
        "pdf_footer": "WebCorp Autopilot · généré après approbation humaine · ",
        "pdf_h1": "Demande de remédiation opérationnelle",
        "pdf_dirigida": "adressée à",
        "pdf_caso": "Cas détecté",
        "titulo": "Titre", "zona": "Zone", "mensajeria": "Transporteur",
        "muestra": "Échantillon (n)", "cruda": "Taux brut",
        "wilson": "Efficacité Wilson LCB 95%", "snapshot": "Instantané des données",
        "no_entregadas": "Commandes non livrées détectées",
        "accion": "Action suggérée", "trazabilidad": "Traçabilité",
        "pdf_meta": ("Méthodologie : classement par la borne inférieure de l'intervalle de Wilson (95%), "
                     "jamais par le taux brut · Ce document a été généré automatiquement par "
                     "WebCorp Autopilot <b>après approbation humaine explicite</b> ; chaque chiffre "
                     "provient d'outils déterministes, pas d'un modèle de langage."),
        "xl_hoja": "Commandes non livrées",
        "xl_cols": ["Commande", "Téléphone", "Municipalité", "Département", "Transporteur",
                    "Statut", "Sous-statut", "COD", "Date"],
        "cc_asunto": "[Autopilot] Suivi call center — {n} commandes non livrées · {muni} × {carrier}",
        "cc_extra": ("Ci-joint l'Excel avec {n} commandes non livrées "
                     "(commande, téléphone, statut, COD) pour le suivi de récupération."),
        "carrier_asunto": "[Autopilot] Remédiation requise — {carrier} à {muni}, {depto}",
        "carrier_extra": "Ci-joint le PDF avec le détail du cas et la méthodologie.",
        "sin_fila": "aucune ligne dans l'estimateur",
        "saludo": "Chère équipe,",
        "intro": "Le système de surveillance de WebCorp a détecté — et un opérateur humain a approuvé d'escalader — le cas suivant :",
        "datos_caso": "Données du cas (instantané {snapshot}) :",
        "evidencia": "Preuve",
        "firma": ("Brouillon généré par WebCorp Autopilot après approbation humaine.\n"
                  "Cet e-mail n'est PAS envoyé automatiquement : il exige une revue et un envoi manuel."),
        "para": "À", "asunto_lbl": "Objet",
        "nota_final": ("Artefacts déterministes générés après approbation humaine. "
                       "Les e-mails sont des BROUILLONS : jamais envoyés automatiquement."),
        "accion_default": "Revue opérationnelle de la paire municipalité-transporteur.",
    },
}


def _t(idioma: str) -> dict:
    return _L.get((idioma or "es").lower()[:2], _L["es"])


# ------------------------------------------------------------------
# Plantillas (deterministicas — solo se llenan espacios con datos
# ya calculados por las tools)
# ------------------------------------------------------------------

def _css_marca(L: dict) -> str:
    return """
  @page { size: letter; margin: 2cm;
    @bottom-center { content: \"""" + L["pdf_footer"] + """\" string(origen);
                     font-size: 8pt; color: #64748B; } }
  body { font-family: Helvetica, Arial, sans-serif; color: #0F172A; font-size: 11pt; }
  .header { background: linear-gradient(135deg, #2563EB 0%, #14B8A6 100%);
            color: #fff; padding: 18px 22px; border-radius: 10px; }
  .header h1 { margin: 0; font-size: 17pt; }
  .header p { margin: 4px 0 0; font-size: 9.5pt; opacity: .9; }
  h2 { font-size: 12pt; color: #1D4ED8; margin: 22px 0 8px; }
  table { border-collapse: collapse; width: 100%; font-size: 10pt; }
  th { text-align: left; background: #F1F5F9; color: #334155; padding: 7px 10px;
       border-bottom: 2px solid #CBD5E1; }
  td { padding: 6px 10px; border-bottom: 1px solid #E2E8F0; }
  .rojo { color: #B91C1C; font-weight: bold; }
  .accion { background: #FEF3C7; border-left: 4px solid #F59E0B; padding: 12px 15px;
            border-radius: 6px; margin-top: 8px; }
  .meta { font-size: 8.5pt; color: #64748B; margin-top: 26px;
          border-top: 1px solid #E2E8F0; padding-top: 8px; }
"""


# Compat: optimizador.py importa esta constante para el CSS de su PDF (es).
_CSS_MARCA = _css_marca(_L["es"])


def _html_pdf(datos: dict, L: dict) -> str:
    c, f = datos["caso"], datos.get("wilson") or {}
    filas_wilson = ""
    if f:
        filas_wilson = f"""
        <tr><th>{L['muestra']}</th><td>{f.get('n', '—')}</td></tr>
        <tr><th>{L['cruda']}</th><td>{f.get('tasa_cruda_pct', '—')}%</td></tr>
        <tr><th>{L['wilson']}</th><td class="rojo">{f.get('wilson_lcb_pct', '—')}%</td></tr>"""
    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"><style>{_css_marca(L)}</style></head>
<body>
  <div class="header">
    <h1>{L['pdf_h1']}</h1>
    <p>WebCorp Autopilot · {L['pdf_dirigida']}: {datos['carrier']} · {datos['fecha_generacion']}</p>
  </div>
  <h2>{L['pdf_caso']}</h2>
  <table>
    <tr><th style="width:40%">{L['titulo']}</th><td>{datos['asunto']}</td></tr>
    <tr><th>{L['zona']}</th><td>{c.get('municipio', '—')}, {c.get('depto', '—')} ({c.get('pais', 'GT')})</td></tr>
    <tr><th>{L['mensajeria']}</th><td>{datos['carrier']}</td></tr>
    {filas_wilson}
    <tr><th>{L['snapshot']}</th><td>{datos['snapshot']}</td></tr>
    <tr><th>{L['no_entregadas']}</th><td>{datos['n_ordenes']}</td></tr>
  </table>
  <h2>{L['accion']}</h2>
  <div class="accion">{datos['accion']}</div>
  <div class="meta">
    {L['trazabilidad']}: {datos['origen_id']} · {L['pdf_meta']}
  </div>
</body></html>"""


def _generar_pdf(datos: dict, base: str, outdir: str, L: dict) -> dict:
    html = _html_pdf(datos, L)
    try:
        from weasyprint import HTML  # import perezoso: si falta pango, caemos a HTML
        ruta = os.path.join(outdir, base + ".pdf")
        HTML(string=html).write_pdf(ruta)
        return {"tipo": "pdf", "archivo": os.path.basename(ruta)}
    except Exception as e:
        ruta = os.path.join(outdir, base + ".html")
        with open(ruta, "w", encoding="utf-8") as fh:
            fh.write(html)
        return {"tipo": "html_imprimible", "archivo": os.path.basename(ruta),
                "nota": f"weasyprint no disponible ({type(e).__name__}); HTML listo para imprimir"}


def _generar_excel(datos: dict, base: str, outdir: str, L: dict) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = L["xl_hoja"][:31]
    c = datos["caso"]
    ws.append([f"WebCorp Autopilot — {datos['asunto']}"])
    ws.append([f"{L['zona']}: {c.get('municipio','—')}, {c.get('depto','—')} · Carrier: {datos['carrier']}"
               f" · {L['snapshot']}: {datos['snapshot']} · {L['trazabilidad']}: {datos['origen_id']}"])
    ws.append([])
    encabezado = L["xl_cols"]
    ws.append(encabezado)
    fila_h = ws.max_row
    for col in range(1, len(encabezado) + 1):
        cell = ws.cell(row=fila_h, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for o in datos["ordenes"]:
        ws.append([o.get("orden"), o.get("telefono"), o.get("municipio"), o.get("depto"),
                   o.get("carrier"), o.get("estatus"), o.get("subestatus"),
                   o.get("cod"), o.get("fecha")])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 13
    for col in "CDEFG":
        ws.column_dimensions[col].width = 20
    ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="1D4ED8")

    ruta = os.path.join(outdir, base + ".xlsx")
    wb.save(ruta)
    return {"tipo": "excel", "archivo": os.path.basename(ruta),
            "ordenes_incluidas": len(datos["ordenes"])}


def _borrador_correo(datos: dict, area: str, base: str, outdir: str, L: dict) -> dict:
    """area: 'mensajeria' | 'callcenter' — las dos unicas soportadas."""
    c, f = datos["caso"], datos.get("wilson") or {}
    if area == "callcenter":
        para = EMAIL_CALLCENTER
        asunto = L["cc_asunto"].format(n=datos["n_ordenes"], muni=c.get("municipio", ""),
                                       carrier=datos["carrier"])
        extra = L["cc_extra"].format(n=datos["n_ordenes"])
    else:
        para = EMAIL_CARRIER_FMT.format(carrier=_slug(datos["carrier"]) or "carrier")
        asunto = L["carrier_asunto"].format(carrier=datos["carrier"], muni=c.get("municipio", ""),
                                            depto=c.get("depto", ""))
        extra = L["carrier_extra"]

    wilson_txt = (f"n={f.get('n','—')} · Wilson LCB 95%: {f.get('wilson_lcb_pct','—')}% "
                  f"({L['cruda'].lower()} {f.get('tasa_cruda_pct','—')}%)") if f else L["sin_fila"]
    cuerpo = f"""{L['saludo']}

{L['intro']}

{datos['asunto']}

{L['datos_caso'].format(snapshot=datos['snapshot'])}
- {L['zona']}: {c.get('municipio','—')}, {c.get('depto','—')} ({c.get('pais','GT')})
- {L['mensajeria']}: {datos['carrier']}
- {L['evidencia']}: {wilson_txt}
- {L['no_entregadas']}: {datos['n_ordenes']}

{L['accion']}: {datos['accion']}

{extra}

{L['trazabilidad']}: {datos['origen_id']}
--
{L['firma']}
"""
    ruta = os.path.join(outdir, f"{base}_correo_{area}.txt")
    with open(ruta, "w", encoding="utf-8") as fh:
        fh.write(f"{L['para']}: {para}\n{L['asunto_lbl']}: {asunto}\n\n{cuerpo}")
    mailto = f"mailto:{para}?subject={urllib.parse.quote(asunto)}&body={urllib.parse.quote(cuerpo)}"
    return {"tipo": "correo_borrador", "area": area, "para": para,
            "archivo": os.path.basename(ruta), "mailto": mailto}


# ------------------------------------------------------------------
# Punto de entrada — lo llama notificar_operaciones DESPUES de la
# aprobacion humana. Recibe todo ya calculado; aqui solo se renderiza.
# `idioma` viene del panel del humano (es|en|fr), no del LLM.
# ------------------------------------------------------------------
def generar(asunto: str, accion: str, origen_id: str, caso: dict,
            wilson: dict = None, snapshot: str = "desconocido",
            ordenes: list = None, idioma: str = "es") -> dict:
    caso = caso or {}
    L = _t(idioma)
    carrier = (caso.get("carrier") or "").upper()
    if not (caso.get("municipio") and carrier):
        return {"generado": False,
                "razon": "sin caso estructurado (municipio+carrier); no aplica artefacto"}

    ordenes = ordenes or []
    outdir = _dir_artefactos()
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    base = f"remediacion_{_slug(caso['municipio'])}_{_slug(carrier)}_{ts}"
    datos = {
        "asunto": asunto, "accion": accion or L["accion_default"],
        "origen_id": origen_id, "caso": caso, "carrier": carrier, "wilson": wilson,
        "snapshot": snapshot, "ordenes": ordenes, "n_ordenes": len(ordenes),
        "fecha_generacion": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    }

    artefactos = []
    # Regla 1 (codigo, no LLM): caso puntual -> PDF dirigido a la mensajeria.
    artefactos.append(_generar_pdf(datos, base, outdir, L))
    correos = [_borrador_correo(datos, "mensajeria", base, outdir, L)]
    # Regla 2: hay clientes afectados -> Excel para call center.
    if ordenes:
        artefactos.append(_generar_excel(datos, base, outdir, L))
        correos.append(_borrador_correo(datos, "callcenter", base, outdir, L))

    return {
        "generado": True,
        "idioma": (idioma or "es").lower()[:2],
        "artefactos": artefactos,
        "correos": correos,
        "nota": L["nota_final"],
    }

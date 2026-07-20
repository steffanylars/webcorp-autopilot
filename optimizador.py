"""
============================================================
WEBCORP AUTOPILOT — Optimizador de asignación zona-carrier
============================================================
¿Qué carrier debería servir cada zona para maximizar las entregas
esperadas? Formulado como problema de asignación generalizada:

  max  Σ_z Σ_c  v_z · e_zc · x_zc
  s.a. Σ_c x_zc = 1                      (cada zona, un carrier)
       Σ_z v_z · x_zc ≤ (1+γ)·V_c        (capacidad: nadie absorbe
                                          volumen infinito de golpe)
       x_zc ∈ {0,1}

donde e_zc es el LÍMITE INFERIOR DE WILSON (95%) — la misma vara
conservadora del resto del sistema, nunca la tasa cruda — y solo
participan pares con evidencia real (n ≥ min_n). Zonas donde un solo
carrier tiene evidencia no entran: no hay decisión que tomar ahí sin
inventar datos.

Método: MILP exacto con PuLP/CBC. Si el solver no está disponible,
cae a una metaheurística (greedy por mejor LCB + reparación de
capacidad + mejora por intercambios 2-opt) y LO DECLARA en 'metodo'.

Todo es determinístico: el LLM puede pedir este plan, pero no
participa en calcularlo.
"""
import os
import csv
import json
from datetime import datetime, timezone

ESTIMADOR_CSV = os.getenv(
    "ESTIMADOR_CSV_PATH",
    os.path.join(os.path.dirname(__file__), "estimador_municipio_mensajeria.csv"),
)
# Ticket promedio por orden ENTREGADA CON COD, USD (query real a la DB,
# jul 2026). Solo órdenes con COD registrado — aproximación declarada.
TICKET_USD = {"GT": 91.18, "SV": 111.66, "HND": 108.47,
              "PAN": 125.51, "CR": 118.32, "NIC": 112.01}

SUPUESTOS = [
    "Efectividades = límite inferior de Wilson 95% (conservador), solo pares con n≥min_n.",
    "Volumen de cada zona = órdenes observadas en la ventana del snapshot (~90 días).",
    "Capacidad: ningún carrier crece más de γ (margen_capacidad) sobre su volumen actual en las zonas optimizables.",
    "Se asume que la efectividad histórica del carrier en la zona se mantiene al recibir el volumen (sin efectos de congestión).",
    "El valor en USD usa el ticket promedio de órdenes COD entregadas por país — no todo el volumen es COD.",
    "Es una recomendación para revisión humana, no una orden de cambio.",
]


def _cargar_zonas(min_n: int, paises: set = None):
    with open(ESTIMADOR_CSV, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    zonas = {}
    for r in rows:
        pais = r["pais"].strip().upper()
        if paises and pais not in paises:
            continue
        if int(r["n"]) < min_n:
            continue
        clave = (pais, r["depto"].strip(), r["municipio"].strip())
        zonas.setdefault(clave, {})[r["mensajeria"].strip().upper()] = {
            "n": int(r["n"]), "lcb": float(r["wilson_lcb"]), "cruda": float(r["tasa_cruda"]),
        }
    # Solo zonas con decisión real: 2+ carriers con evidencia
    return {z: cs for z, cs in zonas.items() if len(cs) >= 2}


def _resolver_milp(zonas, capacidad, margen):
    import pulp
    prob = pulp.LpProblem("asignacion_zona_carrier", pulp.LpMaximize)
    x = {}
    for z, cs in zonas.items():
        for c in cs:
            x[(z, c)] = pulp.LpVariable(f"x_{hash((z, c)) & 0xFFFFFFFF}", cat="Binary")
    # objetivo: entregas esperadas (volumen total de la zona al carrier elegido)
    vol = {z: sum(d["n"] for d in cs.values()) for z, cs in zonas.items()}
    prob += pulp.lpSum(vol[z] * (cs[c]["lcb"] / 100.0) * x[(z, c)]
                       for z, cs in zonas.items() for c in cs)
    for z, cs in zonas.items():
        prob += pulp.lpSum(x[(z, c)] for c in cs) == 1
    for c, cap in capacidad.items():
        prob += pulp.lpSum(vol[z] * x[(z, cc)] for z, cs in zonas.items()
                           for cc in cs if cc == c) <= cap * (1 + margen)
    prob.solve(pulp.PULP_CBC_CMD(msg=0, timeLimit=30))
    if pulp.LpStatus[prob.status] != "Optimal":
        return None, pulp.LpStatus[prob.status]
    asignacion = {}
    for z, cs in zonas.items():
        for c in cs:
            if x[(z, c)].value() and x[(z, c)].value() > 0.5:
                asignacion[z] = c
    return asignacion, "Optimal"


def _resolver_metaheuristica(zonas, capacidad, margen):
    """Greedy por mejor LCB + reparación de capacidad + 2-opt swaps."""
    vol = {z: sum(d["n"] for d in cs.values()) for z, cs in zonas.items()}
    cap = {c: v * (1 + margen) for c, v in capacidad.items()}
    asig = {z: max(cs, key=lambda c: cs[c]["lcb"]) for z, cs in zonas.items()}

    def carga(a):
        out = {c: 0 for c in cap}
        for z, c in a.items():
            out[c] = out.get(c, 0) + vol[z]
        return out

    def objetivo(a):
        return sum(vol[z] * zonas[z][a[z]]["lcb"] / 100.0 for z in a)

    # reparación: mover las zonas de menor pérdida fuera de carriers saturados
    for _ in range(500):
        cg = carga(asig)
        exceso = [c for c in cap if cg.get(c, 0) > cap[c]]
        if not exceso:
            break
        c_mal = exceso[0]
        candidatos = []
        for z, c in asig.items():
            if c != c_mal:
                continue
            for alt in zonas[z]:
                if alt != c_mal and cg.get(alt, 0) + vol[z] <= cap.get(alt, 0):
                    perdida = (zonas[z][c]["lcb"] - zonas[z][alt]["lcb"]) * vol[z]
                    candidatos.append((perdida, z, alt))
        if not candidatos:
            break
        _, z, alt = min(candidatos)
        asig[z] = alt
    # mejora local: reasignaciones simples mientras aporten
    mejoro = True
    while mejoro:
        mejoro = False
        cg = carga(asig)
        for z, cs in zonas.items():
            actual = asig[z]
            for alt in cs:
                if alt == actual:
                    continue
                if cg.get(alt, 0) + vol[z] > cap.get(alt, 0):
                    continue
                if cs[alt]["lcb"] > cs[actual]["lcb"]:
                    asig[z] = alt
                    cg[actual] -= vol[z]
                    cg[alt] = cg.get(alt, 0) + vol[z]
                    mejoro = True
                    actual = alt
    return asig, "Heuristic"


def optimizar(min_n: int = 30, margen_capacidad: float = 0.25,
              paises: list = None, top: int = 15,
              generar_artefactos: bool = True) -> dict:
    """
    Corre la optimización y devuelve el plan de reasignación.
    Solo recomienda: nada se ejecuta sin humanos.
    """
    zonas = _cargar_zonas(min_n, set(p.upper() for p in paises) if paises else None)
    if not zonas:
        return {"ok": False, "razon": f"No hay zonas con 2+ carriers con n≥{min_n}."}

    vol = {z: sum(d["n"] for d in cs.values()) for z, cs in zonas.items()}
    # capacidad actual de cada carrier DENTRO del universo optimizable
    capacidad = {}
    for z, cs in zonas.items():
        for c, d in cs.items():
            capacidad[c] = capacidad.get(c, 0) + d["n"]

    # línea base: la mezcla actual (cada carrier conserva sus n)
    base_entregas = sum(d["n"] * d["lcb"] / 100.0 for cs in zonas.values() for d in cs.values())

    metodo = "MILP exacto (PuLP + CBC)"
    try:
        asignacion, status = _resolver_milp(zonas, capacidad, margen_capacidad)
        if asignacion is None:
            raise RuntimeError(f"solver status: {status}")
    except Exception as e:
        metodo = f"Metaheurística greedy + 2-opt (MILP no disponible: {type(e).__name__})"
        asignacion, status = _resolver_metaheuristica(zonas, capacidad, margen_capacidad)

    opt_entregas = sum(vol[z] * zonas[z][asignacion[z]]["lcb"] / 100.0 for z in asignacion)
    ganancia = opt_entregas - base_entregas

    recomendaciones = []
    ganancia_usd = 0.0
    for z, elegido in asignacion.items():
        pais, depto, municipio = z
        cs = zonas[z]
        dominante = max(cs, key=lambda c: cs[c]["n"])   # quién lleva hoy el volumen
        if elegido == dominante:
            continue
        delta = vol[z] * (cs[elegido]["lcb"] - cs[dominante]["lcb"]) / 100.0
        usd = delta * TICKET_USD.get(pais, 100.0)
        ganancia_usd += max(usd, 0)
        recomendaciones.append({
            "pais": pais, "depto": depto, "municipio": municipio,
            "volumen": vol[z],
            "carrier_actual": dominante, "ef_actual_lcb": cs[dominante]["lcb"],
            "carrier_recomendado": elegido, "ef_recomendada_lcb": cs[elegido]["lcb"],
            "entregas_extra_estimadas": round(delta, 1),
            "usd_estimados": round(usd, 0),
        })
    recomendaciones.sort(key=lambda r: -r["entregas_extra_estimadas"])

    resultado = {
        "ok": True,
        "metodo": metodo,
        "status_solver": status,
        "parametros": {"min_n": min_n, "margen_capacidad": margen_capacidad},
        "universo": {
            "zonas_optimizables": len(zonas),
            "carriers": sorted(capacidad),
            "volumen_total": sum(vol.values()),
            "nota": "Solo zonas con 2+ carriers con evidencia (n≥min_n); el resto no entra — sin evidencia no hay decisión.",
        },
        "resumen": {
            "entregas_esperadas_actual": round(base_entregas, 0),
            "entregas_esperadas_optimo": round(opt_entregas, 0),
            "entregas_extra": round(ganancia, 0),
            "mejora_pct": round(100.0 * ganancia / base_entregas, 2) if base_entregas else 0,
            "usd_estimados": round(ganancia_usd, 0),
            "zonas_con_cambio": len(recomendaciones),
            "ventana": "~90 días (ventana del snapshot del estimador)",
        },
        "recomendaciones": recomendaciones[:top],
        "supuestos": SUPUESTOS,
    }
    if generar_artefactos:
        try:
            resultado["artefactos"] = _artefactos(resultado, recomendaciones)
        except Exception as e:
            resultado["artefactos"] = {"error": f"{type(e).__name__}: {e}"}
    return resultado


# ------------------------------------------------------------------
# Descargables: Excel con el plan completo + PDF resumen con marca
# ------------------------------------------------------------------
def _dir_artefactos():
    d = os.getenv("ARTEFACTOS_DIR", os.path.join(os.path.dirname(__file__), "artefactos"))
    os.makedirs(d, exist_ok=True)
    return d


def _artefactos(res, recomendaciones_completas):
    outdir = _dir_artefactos()
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    base = f"plan_optimizacion_{ts}"
    out = []

    # Excel: plan completo
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan de reasignacion"
    r = res["resumen"]
    ws.append(["WebCorp Autopilot — Plan de optimización de asignación zona-carrier"])
    ws.append([f"Método: {res['metodo']} · Zonas: {res['universo']['zonas_optimizables']}"
               f" · Entregas extra estimadas: {r['entregas_extra']:.0f} (+{r['mejora_pct']}%)"
               f" · ~${r['usd_estimados']:,.0f} USD COD · Ventana: {r['ventana']}"])
    ws.append([])
    head = ["País", "Depto", "Municipio", "Volumen (n)", "Carrier actual", "Ef. actual (Wilson LCB %)",
            "Carrier recomendado", "Ef. recomendada (Wilson LCB %)", "Entregas extra", "USD estimados"]
    ws.append(head)
    fila_h = ws.max_row
    for col in range(1, len(head) + 1):
        cell = ws.cell(row=fila_h, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for rec in recomendaciones_completas:
        ws.append([rec["pais"], rec["depto"], rec["municipio"], rec["volumen"],
                   rec["carrier_actual"], rec["ef_actual_lcb"],
                   rec["carrier_recomendado"], rec["ef_recomendada_lcb"],
                   rec["entregas_extra_estimadas"], rec["usd_estimados"]])
    ws.append([])
    ws.append(["Supuestos y límites:"])
    for sup in res["supuestos"]:
        ws.append([sup])
    for col, ancho in zip("ABCDEFGHIJ", (7, 16, 22, 12, 15, 22, 19, 26, 14, 14)):
        ws.column_dimensions[col].width = ancho
    ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="1D4ED8")
    ruta = os.path.join(outdir, base + ".xlsx")
    wb.save(ruta)
    out.append({"tipo": "excel", "archivo": os.path.basename(ruta)})

    # PDF resumen (misma plantilla con marca que la remediación)
    filas = "".join(
        f"<tr><td>{x['municipio']}, {x['depto']} ({x['pais']})</td><td>{x['carrier_actual']} · {x['ef_actual_lcb']}%</td>"
        f"<td><b>{x['carrier_recomendado']} · {x['ef_recomendada_lcb']}%</b></td>"
        f"<td>{x['entregas_extra_estimadas']}</td><td>${x['usd_estimados']:,.0f}</td></tr>"
        for x in res["recomendaciones"][:12])
    sup = "".join(f"<li>{sx}</li>" for sx in res["supuestos"])
    from remediacion import _CSS_MARCA
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><style>{_CSS_MARCA}</style></head><body>
    <div class="header"><h1>Plan de optimización de asignación</h1>
      <p>WebCorp Autopilot · {res['metodo']} · generado {ts} UTC</p></div>
    <h2>Resumen</h2>
    <table>
      <tr><th>Zonas optimizables</th><td>{res['universo']['zonas_optimizables']}</td></tr>
      <tr><th>Entregas extra estimadas</th><td class="rojo">{r['entregas_extra']:.0f} (+{r['mejora_pct']}%) en {r['ventana']}</td></tr>
      <tr><th>Valor COD estimado</th><td>${r['usd_estimados']:,.0f} USD</td></tr>
      <tr><th>Zonas con cambio recomendado</th><td>{r['zonas_con_cambio']}</td></tr>
    </table>
    <h2>Top cambios recomendados</h2>
    <table><tr><th>Zona</th><th>Hoy</th><th>Recomendado</th><th>Entregas extra</th><th>USD</th></tr>{filas}</table>
    <h2>Supuestos y límites</h2><div class="accion"><ul>{sup}</ul></div>
    <div class="meta">Optimización determinística sobre el límite inferior de Wilson (95%). Recomendación para
    revisión humana — el Autopilot no ejecuta cambios de asignación.</div>
    </body></html>"""
    try:
        from weasyprint import HTML
        ruta = os.path.join(outdir, base + ".pdf")
        HTML(string=html).write_pdf(ruta)
        out.append({"tipo": "pdf", "archivo": os.path.basename(ruta)})
    except Exception:
        ruta = os.path.join(outdir, base + ".html")
        with open(ruta, "w", encoding="utf-8") as fh:
            fh.write(html)
        out.append({"tipo": "html_imprimible", "archivo": os.path.basename(ruta)})
    return out


if __name__ == "__main__":
    res = optimizar()
    print(json.dumps({k: v for k, v in res.items() if k != "recomendaciones"},
                     ensure_ascii=False, indent=2))
    for rec in res.get("recomendaciones", [])[:5]:
        print(rec)
